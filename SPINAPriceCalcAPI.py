# This is the main code
# Author: M. Poul
# Created: 01/2018
#
#
from openpyxl import Workbook


def create_workbook():
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = 42

    # Rows can also be appended
    ws.append([1, 2, 3])

    # Python types will automatically be converted
    import datetime
    ws['A2'] = datetime.datetime.now()

    # Save the file
    wb.save("sample.xlsx")

#This should open excel application
#Then save the workbook
#Then close the workbook
def recalculate_excel_win(path_to_excel):
    import win32com.client as win32

    excel = win32.gencache.EnsureDispatch('Excel.Application')
    workbook = excel.Workbooks.Open(path_to_excel)
    # this must be the absolute path (r'C:/abc/def/ghi')
    workbook.Save()
    workbook.Close()
    excel.Quit()

import argparse

parser = argparse.ArgumentParser(description='SPINA Price Calc Generator Python API.')
parser.add_argument('cpu', metavar='c', type=int, help='Expected CPU in virtual cores')
parser.add_argument('ram', metavar='r', type=int, help='Expected RAM in Gb')
parser.add_argument('flash', metavar='sf', type=int, help='Expected Flash Storage Size in Gb')
parser.add_argument('slow', metavar='ss', type=int, help='Expected NonFlash Storages Size in Gb')
parser.add_argument('gpfs', metavar='sg', type=int, help='Expected GPFS (IBM Clustered Filesystem) Size in Gb')
parser.add_argument('bck', metavar='b', type=int, help='Expected Backup Size in Gb')
parser.add_argument('--sum', dest='accumulate', action='store_const',
                    const=sum, default=max,
                    help='sum the integers (default: find the max)')

args = parser.parse_args()
print(args.accumulate(args.integers))
